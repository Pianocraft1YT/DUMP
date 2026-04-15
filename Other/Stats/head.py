#!/usr/bin/env python3
"""
Add player‑based color coding to the stats Excel file (no images).

Reads 'stats_output.xlsx' (or a given input file) and creates a new file
where every row belonging to the same player has a unique pastel background
color. The color is derived from the player's name/UUID and remains
consistent across runs.

Usage:
    python color_stats.py [input_file] [output_file]

If input_file is omitted, defaults to 'stats_output.xlsx'.
If output_file is omitted, defaults to 'stats_colored.xlsx'.
"""

import sys
import hashlib
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

def get_player_color(player):
    """Generate a pastel RGB color from a player string."""
    hash_val = int(hashlib.md5(player.encode()).hexdigest()[:6], 16)
    # Pastel colors: mix with white (0xC0,0xC0,0xC0) -> (r+0xC0)//2 etc.
    r = (hash_val >> 16) & 0xFF
    g = (hash_val >> 8) & 0xFF
    b = hash_val & 0xFF
    # Lighten
    r = (r + 0xC0) // 2
    g = (g + 0xC0) // 2
    b = (b + 0xC0) // 2
    return f"{r:02X}{g:02X}{b:02X}"

def main():
    # File names
    input_file = sys.argv[1] if len(sys.argv) > 1 else "stats_output.xlsx"
    output_file = sys.argv[2] if len(sys.argv) > 2 else "stats_colored.xlsx"

    if not Path(input_file).exists():
        print(f"Error: input file '{input_file}' not found.", file=sys.stderr)
        sys.exit(1)

    print(f"Loading {input_file}...")
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    # Find the Player column (assuming headers in first row)
    headers = [cell.value for cell in ws[1]]
    try:
        player_col = headers.index("Player") + 1   # 1‑based column index
    except ValueError:
        print("Error: could not find a 'Player' column.", file=sys.stderr)
        sys.exit(1)

    # Process each data row (skip header)
    print("Applying colors...")
    for row in range(2, ws.max_row + 1):
        player = ws.cell(row=row, column=player_col).value
        if not player:
            continue

        color = get_player_color(player)
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # Apply fill to every cell in this row
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill

    # Save the new workbook
    wb.save(output_file)
    print(f"Done. Created '{output_file}' with color‑coded rows.")

if __name__ == "__main__":
    main()
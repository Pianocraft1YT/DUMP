from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "LACP Modes"

# ---------- Styles ----------
title_font = Font(size=14, bold=True, color="FFFFFF")
header_font = Font(bold=True, color="FFFFFF")
body_font = Font(color="FFFFFF")

title_fill = PatternFill("solid", fgColor="1B1F23")
header_fill = PatternFill("solid", fgColor="5B6670")
row_fill = PatternFill("solid", fgColor="23272B")

align_left = Alignment(horizontal="left", vertical="center")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ---------- Title ----------
ws.merge_cells("A1:C1")
title = ws["A1"]
title.value = "LACP Mode Combinations and Channel Establishment"
title.font = title_font
title.fill = title_fill
title.alignment = align_left

# ---------- Headers ----------
headers = ["S1", "S2", "Channel Establishment"]
for col, text in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=col, value=text)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = align_left
    cell.border = thin_border

# ---------- Data ----------
data = [
    ("On", "On", "Yes"),
    ("On", "Active/Passive", "No"),
    ("Active", "Active", "Yes"),
    ("Active", "Passive", "Yes"),
    ("Passive", "Active", "Yes"),
    ("Passive", "Passive", "No"),
]

for row_idx, row in enumerate(data, start=3):
    for col_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = body_font
        cell.fill = row_fill
        cell.alignment = align_left
        cell.border = thin_border

# ---------- Column Widths ----------
ws.column_dimensions["A"].width = 16
ws.column_dimensions["B"].width = 26
ws.column_dimensions["C"].width = 32

# ---------- Freeze Header ----------
ws.freeze_panes = "A3"

# ---------- Save ----------
wb.save("LACP_modes_formatted.xlsx")
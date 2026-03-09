#!/usr/bin/env python3
"""
Convert the output of the first parser (lines: "<stat_key> <player> <value>")
into a nicely formatted Excel spreadsheet with color formatting.

Features:
- Auto-detects UTF-8 or UTF-16 encoding when reading from a file.
- Creates an Excel file with bold headers, frozen top row, autofilter,
  auto-adjusted column widths, and number formatting.
- Applies a two‑color scale (green high, red low) to the Value column.
- Handles large datasets efficiently.

Usage:
    python create_excel.py [input_file] [-o output_file]

If input_file is omitted, reads from stdin (assumes UTF-8).
"""

import argparse
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

def read_file_with_encoding_detection(path):
    """
    Attempt to read the file using UTF-8, then UTF-16, then system default.
    Returns the file content as a string.
    """
    encodings = ['utf-8', 'utf-16', sys.getdefaultencoding()]
    for enc in encodings:
        try:
            return path.read_text(encoding=enc)
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError(f"Could not decode {path} with any of {encodings}")

def read_input_lines(source):
    """Yield non-empty lines from a file path or stdin."""
    if source is sys.stdin:
        # stdin: assume UTF-8 (common for pipes)
        for line in source:
            line = line.strip()
            if line:
                yield line
    else:
        content = read_file_with_encoding_detection(Path(source))
        for line in content.splitlines():
            line = line.strip()
            if line:
                yield line

def main():
    parser = argparse.ArgumentParser(description="Create color‑formatted Excel spreadsheet from stats data.")
    parser.add_argument('input', nargs='?', help='Input file (if omitted, reads from stdin)')
    parser.add_argument('-o', '--output', default='stats_output.xlsx',
                        help='Output Excel file (default: stats_output.xlsx)')
    args = parser.parse_args()

    # Determine input source
    if args.input:
        input_path = Path(args.input)
        if not input_path.exists():
            print(f"Error: Input file '{args.input}' not found.", file=sys.stderr)
            sys.exit(1)
        lines = read_input_lines(args.input)
    else:
        lines = read_input_lines(sys.stdin)

    # Create workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Stats"

    # Write header
    headers = ["Stat Key", "Player", "Value"]
    ws.append(headers)

    # Process data rows
    row_count = 0
    for line in lines:
        parts = line.split()
        if len(parts) < 3:
            print(f"Skipping malformed line: {line}", file=sys.stderr)
            continue
        stat_key = parts[0]
        player = parts[1]
        try:
            value = int(parts[2])
        except ValueError:
            print(f"Invalid value in line: {line}", file=sys.stderr)
            continue
        ws.append([stat_key, player, value])
        row_count += 1

    if row_count == 0:
        print("No valid data found.", file=sys.stderr)
        wb.save(args.output)
        print(f"Empty workbook saved as {args.output}")
        return

    # --- Apply formatting ---

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Autofilter on all columns
    ws.auto_filter.ref = ws.dimensions

    # Style for header row
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Auto-adjust column widths (capped at 50)
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Format the Value column as numbers with thousand separator
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = '#,##0'

    # --- Color formatting: two‑color scale on Value column (C2:C<last>) ---
    last_row = ws.max_row
    value_range = f'C2:C{last_row}'
    # Green for highest, red for lowest
    color_scale = ColorScaleRule(
        start_type='min', start_color='FF63BE7B',  # green
        end_type='max', end_color='FFFA8072'       # salmon/red
    )
    ws.conditional_formatting.add(value_range, color_scale)

    # Save workbook
    wb.save(args.output)
    print(f"Excel file saved as {args.output} with {row_count} rows.")

if __name__ == "__main__":
    main()